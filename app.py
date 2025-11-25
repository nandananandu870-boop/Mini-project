import os
import time
import tempfile
import traceback
from flask import Flask, session, redirect, url_for, request, render_template, flash, send_file
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import Flow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from rapidfuzz import fuzz
from openpyxl import Workbook

# --- CONFIG ---
os.environ["OAUTHLIB_INSECURE_TRANSPORT"] = "1"  # dev only
APP_SECRET_KEY = "NANDANA_SUPER_SECRET_KEY_123456789"  # replace for production
SCOPES = ["https://www.googleapis.com/auth/gmail.modify"]
CLIENT_SECRETS_FILE = "client_secrets.json"
OAUTH2_CALLBACK = os.environ.get(
    "OAUTH2_CALLBACK",
    "http://localhost:5000/oauth2callback"
)

app = Flask(__name__)
app.secret_key = APP_SECRET_KEY


# --- HELPERS: credentials & Gmail service ---
def creds_to_dict(creds: Credentials):
    return {
        "token": creds.token,
        "refresh_token": creds.refresh_token,
        "token_uri": creds.token_uri,
        "client_id": creds.client_id,
        "client_secret": creds.client_secret,
        "scopes": creds.scopes,
    }


def creds_from_session():
    if "credentials" not in session:
        return None
    return Credentials(**session["credentials"])


def safe_build_service(creds, retries=3, delay=1.0):
    """Build Gmail service and validate by calling getProfile (with minimal retries)."""
    for attempt in range(retries):
        try:
            service = build("gmail", "v1", credentials=creds, cache_discovery=False)
            service.users().getProfile(userId="me").execute()
            return service
        except Exception as ex:
            print(f"[GMAIL CONNECT] attempt {attempt+1}/{retries} failed: {ex}")
            if attempt < retries - 1:
                time.sleep(delay)
            else:
                return None


# --- Duplicate detection ---
def find_near_duplicates(emails, threshold=70):
    """
    Return:
      - pairs: list of {"email1": e1, "email2": e2, "similarity": score}
      - unique_near_ids: set of message ids that appear in any near pair
    """
    pairs = []
    n = len(emails)
    for i in range(n):
        s1 = (emails[i].get("subject") or "").strip()
        if not s1:
            continue
        for j in range(i + 1, n):
            s2 = (emails[j].get("subject") or "").strip()
            if not s2:
                continue
            sim = fuzz.token_sort_ratio(s1, s2)
            if threshold <= sim < 100:
                pairs.append({
                    "email1": emails[i],
                    "email2": emails[j],
                    "similarity": round(sim, 4)
                })
    unique_ids = set()
    for p in pairs:
        unique_ids.add(p["email1"]["id"])
        unique_ids.add(p["email2"]["id"])
    return pairs, unique_ids


# --- Label helpers ---
def get_or_create_label(service, label_name: str):
    labels_resp = service.users().labels().list(userId="me").execute()
    for lbl in labels_resp.get("labels", []):
        if lbl.get("name", "").lower() == label_name.lower():
            return lbl["id"]
    body = {"name": label_name, "labelListVisibility": "labelShow", "messageListVisibility": "show"}
    created = service.users().labels().create(userId="me", body=body).execute()
    return created["id"]


def batch_add_label(service, ids, label_id):
    if not ids:
        return
    body = {"ids": list(ids), "addLabelIds": [label_id], "removeLabelIds": []}
    service.users().messages().batchModify(userId="me", body=body).execute()


def batch_remove_label(service, ids, label_id):
    if not ids:
        return
    body = {"ids": list(ids), "addLabelIds": [], "removeLabelIds": [label_id]}
    service.users().messages().batchModify(userId="me", body=body).execute()


# --- ROUTES ---
@app.route("/")
def index():
    creds = creds_from_session()
    signed_in = bool(creds and creds.valid)
    return render_template("index.html", signed_in=signed_in)


@app.route("/authorize")
def authorize():
    try:
        flow = Flow.from_client_secrets_file(CLIENT_SECRETS_FILE, scopes=SCOPES, redirect_uri=OAUTH2_CALLBACK)
        auth_url, state = flow.authorization_url(access_type="offline", include_granted_scopes="true", prompt="consent")
        session["state"] = state
        return redirect(auth_url)
    except Exception as ex:
        flash("Authorization error. Check client_secrets.json and OAuth setup.")
        print("authorize error:", ex)
        return redirect(url_for("index"))


@app.route("/oauth2callback")
def oauth2callback():
    state = session.get("state")
    try:
        flow = Flow.from_client_secrets_file(CLIENT_SECRETS_FILE, scopes=SCOPES, state=state, redirect_uri=OAUTH2_CALLBACK)
        flow.fetch_token(authorization_response=request.url)
    except Exception as ex:
        # mismatched state or fetch failure
        print("oauth2callback error:", ex)
        flash("Login failed or session expired. Please try again.")
        return redirect(url_for("index"))

    session["credentials"] = creds_to_dict(flow.credentials)
    flash("✅ Signed in with Google.")
    return redirect(url_for("index"))


@app.route("/signout")
def signout():
    session.clear()
    flash("Signed out.")
    return redirect(url_for("index"))


@app.route("/dedupe", methods=["POST"])
def dedupe():
    """Main scan route. Returns results page with counts and tables."""
    creds = creds_from_session()
    if not creds:
        flash("Please sign in with Google first.")
        return redirect(url_for("index"))

    service = safe_build_service(creds)
    if not service:
        flash("Could not connect to Gmail API.")
        return redirect(url_for("index"))

    # How many messages to fetch (user provided)
    try:
        max_emails = int(request.form.get("max_emails", 500))
        max_emails = max(10, min(2000, max_emails))
    except Exception:
        max_emails = 500

    emails = []
    page_token = None
    fetched = 0
    page_size = 300  # metadata fetch per page (tuneable)

    try:
        while fetched < max_emails:
            resp = service.users().messages().list(userId="me", maxResults=min(page_size, max_emails - fetched),
                                                   pageToken=page_token).execute()
            for m in resp.get("messages", []):
                # fetch only metadata headers (faster)
                msg = service.users().messages().get(
                    userId="me", id=m["id"], format="metadata", metadataHeaders=["From", "Subject", "Date"]
                ).execute()
                headers = {h["name"]: h["value"] for h in msg.get("payload", {}).get("headers", [])}
                ts = int(msg.get("internalDate", "0"))
                emails.append({
                    "id": m["id"],
                    "from": headers.get("From", ""),
                    "subject": headers.get("Subject", "") or "",
                    "date": headers.get("Date", ""),
                    "ts": ts
                })
                fetched += 1
                if fetched >= max_emails:
                    break

            page_token = resp.get("nextPageToken")
            if not page_token:
                break
    except HttpError as he:
        print("Gmail fetch error:", he)
        flash("Gmail API error while fetching emails. Check logs.")
        return redirect(url_for("index"))
    except Exception as ex:
        print("Fetch exception:", ex)
        flash("Unexpected error while fetching emails.")
        return redirect(url_for("index"))

    # ---- Exact duplicates groups (sender + subject) ----
    groups = {}
    for e in emails:
        key = (e["from"].strip(), e["subject"].strip())
        groups.setdefault(key, []).append(e)

    duplicate_groups = []
    redundant_ids = []  # these are older copies we mark as duplicates and show/delete
    redundant_emails_for_table = []

    for key, grp in groups.items():
        if len(grp) > 1:
            # sort by timestamp desc (newest first) and keep the first as safe
            grp_sorted = sorted(grp, key=lambda x: x["ts"], reverse=True)
            duplicate_groups.append(grp_sorted)
            # older copies -> redundant
            for e in grp_sorted[1:]:
                redundant_ids.append(e["id"])
                redundant_emails_for_table.append(e)

    # Count exact duplicates as the number of redundant messages (older copies)
    duplicate_count = len(redundant_ids)

    # Apply DUPLICATE label to redundant copies
    try:
        lbl_duplicate_id = get_or_create_label(service, "DUPLICATE")
        if redundant_ids:
            # batch modify in chunks of 1000 (API limit)
            for i in range(0, len(redundant_ids), 1000):
                batch_add_label(service, redundant_ids[i:i + 1000], lbl_duplicate_id)
    except Exception as ex:
        print("Label/DUPLICATE error:", ex)
        flash("Could not apply DUPLICATE label to messages (check scopes).")

    # ---- Near duplicates (subject similarity) ----
    pairs, near_ids = find_near_duplicates(emails, threshold=70)
    # Create unique email list for near duplicates for table
    id_to_email = {e["id"]: e for e in emails}
    near_unique_emails = [id_to_email[_id] for _id in sorted(near_ids, key=lambda i: id_to_email[i]["ts"], reverse=True)]
    near_count = len(near_unique_emails)

    try:
        lbl_near_id = get_or_create_label(service, "NEAR_DUPLICATE")
        if near_unique_emails:
            uid_list = [e["id"] for e in near_unique_emails]
            for i in range(0, len(uid_list), 1000):
                batch_add_label(service, uid_list[i:i + 1000], lbl_near_id)
    except Exception as ex:
        print("Label/NEAR_DUPLICATE error:", ex)
        flash("Could not apply NEAR_DUPLICATE label to messages (check scopes).")

    # Save groups and label id for smart delete later
    session["duplicate_groups"] = duplicate_groups
    session["duplicate_label_id"] = lbl_duplicate_id if 'lbl_duplicate_id' in locals() else None

    # Render results: duplicates shown are the redundant copies (older ones) so counts match
    return render_template(
        "results.html",
        fetched=fetched,
        uniques=max(0, len(emails) - duplicate_count),
        duplicate_count=duplicate_count,
        duplicates=redundant_emails_for_table,
        similars=pairs,
        near_list=near_unique_emails,
        near_count=near_count
    )


@app.route("/smart_delete", methods=["POST"])
def smart_delete():
    creds = creds_from_session()
    if not creds:
        flash("Please sign in first.")
        return redirect(url_for("index"))

    service = safe_build_service(creds)
    if not service:
        flash("Gmail connection failed.")
        return redirect(url_for("index"))

    duplicate_groups = session.get("duplicate_groups", [])
    duplicate_label_id = session.get("duplicate_label_id", None)

    deleted = []
    kept = []

    for grp in duplicate_groups:
        if not grp:
            continue
        keep = grp[0]
        kept.append(keep)
        # remove DUPLICATE label from kept (if label exists)
        if duplicate_label_id:
            try:
                batch_remove_label(service, [keep["id"]], duplicate_label_id)
            except Exception:
                pass
        # delete the older copies
        for e in grp[1:]:
            try:
                service.users().messages().delete(userId="me", id=e["id"]).execute()
                deleted.append(e)
            except HttpError as he:
                print("Smart delete error:", he)
            except Exception as ex:
                print("Smart delete error:", ex)

    flash(f"Smart delete finished. Deleted {len(deleted)} messages.")
    return render_template("deleted.html", kept=kept, deleted=deleted)


@app.route("/delete", methods=["POST"])
def delete_selected():
    """Delete selected ids from table (permanent)."""
    creds = creds_from_session()
    if not creds:
        flash("Sign in first.")
        return redirect(url_for("index"))
    service = safe_build_service(creds)
    if not service:
        flash("Gmail connection failed.")
        return redirect(url_for("index"))

    ids = request.form.getlist("ids")
    deleted = []
    for mid in ids:
        try:
            service.users().messages().delete(userId="me", id=mid).execute()
            deleted.append(mid)
        except Exception as ex:
            print("delete_selected error:", ex)

    flash(f"Deleted {len(deleted)} selected messages.")
    return redirect(url_for("index"))


@app.route("/export_excel")
def export_excel():
    """Export Summary, Exact duplicates, Near duplicates, and Near pairs to multi-sheet XLSX."""
    creds = creds_from_session()
    if not creds:
        flash("Sign in first.")
        return redirect(url_for("index"))
    service = safe_build_service(creds)
    if not service:
        flash("Gmail connection failed.")
        return redirect(url_for("index"))

    # Attempt to reuse session-stored results to avoid re-scanning
    # The dedupe route stores `duplicates` (redundant copies) & `similars` & `near_list` only in template render, so we saved groups earlier in session.
    duplicate_groups = session.get("duplicate_groups", [])
    # Build duplicates list (redundant copies) from groups
    redundant_emails = []
    for grp in duplicate_groups:
        if len(grp) > 1:
            redundant_emails.extend(grp[1:])

    # Near duplicates: we cannot rely on pairs stored in session (not stored), so re-run a lightweight fetch
    # For reliability, re-run dedupe logic in-memory but with smaller limit: use previously scanned messages if present in session?
    # For simplicity, we will re-use the last page fetch by calling /dedupe again would be expensive. Instead, we use what's stored in session (if any)
    # Fallback: create empty sheets if not available.
    # We'll rely on session-stored 'last_scan_emails' if user modified code to store it. For now, export what we have.
    # (In our flow, duplicates were labeled in Gmail, so this is still useful.)

    # Create workbook
    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"

    total_fetched = session.get("last_fetched", None)
    if total_fetched is None:
        # best-effort: count messages labelled DUPLICATE / NEAR_DUPLICATE via Gmail labels (expensive) - skip and leave blank
        total_fetched = ""

    # Summary sheet
    ws_sum.append(["Metric", "Value"])
    ws_sum.append(["Total Emails Fetched (scan)", total_fetched])
    ws_sum.append(["Exact Duplicate Emails (older copies)", len(redundant_emails)])
    # For near duplicates, we can probe labels (fast if labels exist)
    # try to find NEAR_DUPLICATE label and count messages with it
    near_count = ""
    try:
        labels = service.users().labels().list(userId="me").execute().get("labels", [])
        near_label = next((l for l in labels if l.get("name", "").lower() == "near_duplicate"), None)
        if near_label:
            resp = service.users().messages().list(userId="me", labelIds=[near_label["id"]], maxResults=1).execute()
            # API doesn't return totalResults always; but list returns 'resultSizeEstimate'
            near_count = resp.get("resultSizeEstimate", "")
    except Exception:
        near_count = ""

    ws_sum.append(["Near Duplicate Emails (unique flagged)", near_count])

    # Exact duplicates sheet
    ws_dup = wb.create_sheet("ExactDuplicates")
    ws_dup.append(["From", "Subject", "Date", "Message ID", "Timestamp"])
    for e in redundant_emails:
        ws_dup.append([e.get("from"), e.get("subject"), e.get("date"), e.get("id"), e.get("ts")])

    # Near duplicates sheet (try to fetch messages by NEAR_DUPLICATE label)
    ws_near = wb.create_sheet("NearDuplicates")
    ws_near.append(["From", "Subject", "Date", "Message ID"])
    try:
        if near_label:
            # iterate all messages with this label (page through)
            token = None
            while True:
                resp = service.users().messages().list(userId="me", labelIds=[near_label["id"]], pageToken=token,
                                                       maxResults=500).execute()
                for m in resp.get("messages", []):
                    msg = service.users().messages().get(userId="me", id=m["id"], format="metadata",
                                                         metadataHeaders=["From", "Subject", "Date"]).execute()
                    headers = {h["name"]: h["value"] for h in msg.get("payload", {}).get("headers", [])}
                    ws_near.append([headers.get("From", ""), headers.get("Subject", ""), headers.get("Date", ""), m["id"]])
                token = resp.get("nextPageToken")
                if not token:
                    break
    except Exception as ex:
        print("export near fetch error:", ex)

    # Near pairs sheet (optional) - we'll leave empty if we don't have pairs cached
    ws_pairs = wb.create_sheet("NearPairs")
    ws_pairs.append(["From 1", "Subject 1", "From 2", "Subject 2", "Similarity (%)"])

    # Save to temporary file and send
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp_name = tmp.name
    tmp.close()
    wb.save(tmp_name)

    return send_file(tmp_name, as_attachment=True, download_name="gmail_scan_report.xlsx")


# --- error handlers (basic) ---
@app.errorhandler(500)
def internal_error(e):
    # try to show friendly page with stack trace summary (debug off)
    tb = traceback.format_exc()
    print("Internal error:", tb)
    try:
        return render_template("500.html", error=str(e)), 500
    except Exception:
        return "Internal server error. Check logs.", 500


if __name__ == "__main__":
    app.run("0.0.0.0", port=5000, debug=False)
